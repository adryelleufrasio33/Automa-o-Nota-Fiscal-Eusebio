import os
import re
import pdfplumber
from openpyxl import Workbook
from datetime import datetime
import subprocess


directory = r"C:\Users\PC 02\Desktop\arquivos"
soffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"

def norm(s: str) -> str:
    return re.sub(r"[ \t]+", " ", (s or "").replace("\xa0", " ").replace("\u202f", " ")).strip()

def extrair_razao_posicional(pdf_path: str) -> str | None:
    stop_tokens = {
        "E-mail", "Email", "Endereço", "Endereco", "Telefone",
        "CPF/CNPJ", "CPF", "CNPJ", "Insc.", "Inscr.", "Municipal",
        "Estadual", "UF", "Cidade", "C.E.P", "CEP"
    }

    def limpa_val(v: str) -> str:
        v = norm(v)
        v = re.split(r"\s+(E-?mail|Endere[cç]o|Telefone|CPF/?CNPJ|CPF|CNPJ|Insc\.?|Inscr\.?)\b",
                     v, maxsplit=1, flags=re.I)[0]
        return norm(v)

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        words = page.extract_words(x_tolerance=2.5, y_tolerance=4)
        if words:
            rows: dict[int, list] = {}
            for w in words:
                ymid = round((w["top"] + w["bottom"]) / 2)
                rows.setdefault(ymid, []).append(w)

            ys = sorted(rows.keys())
            for idx, y in enumerate(ys):
                linha = sorted(rows[y], key=lambda w: w["x0"])
                for i in range(len(linha) - 1):
                    t1 = linha[i]["text"]
                    t2 = linha[i + 1]["text"]
                    if re.fullmatch(r"Raz[aã]o", t1, flags=re.I) and re.fullmatch(r"Social", t2, flags=re.I):
                        anchor_x = linha[i + 1]["x1"]

                        direita = []
                        for w in linha[i + 2:]:
                            if w["x0"] <= anchor_x:
                                continue
                            if any(re.fullmatch(tok, w["text"], flags=re.I) for tok in stop_tokens):
                                break
                            direita.append(w["text"])
                        val = limpa_val(" ".join(direita))
                        if val:
                            return val

                        if idx + 1 < len(ys):
                            prox = sorted(rows[ys[idx + 1]], key=lambda w: w["x0"])
                            direita2 = []
                            for w in prox:
                                if w["x0"] < anchor_x - 6:
                                    continue
                                if any(re.fullmatch(tok, w["text"], flags=re.I) for tok in stop_tokens):
                                    break
                                direita2.append(w["text"])
                            val2 = limpa_val(" ".join(direita2))
                            if val2:
                                return val2

    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            txt = page.extract_text(x_tolerance=2.5, y_tolerance=4) or ""
    except Exception:
        txt = ""

    if txt:
        m = re.search(r"Raz[aã]o\s+Social\s*[:\-]?\s*(.+)", txt, flags=re.I)
        if m:
            cand = limpa_val(m.group(1))
            if cand:
                return cand

        linhas = [norm(l) for l in txt.splitlines() if norm(l)]
        for i, l in enumerate(linhas):
            if re.search(r"\bRaz[aã]o\s+Social\b", l, flags=re.I):
                pos = re.sub(r".*?\bRaz[aã]o\s+Social\b[:\-]?\s*", "", l, flags=re.I)
                pos = limpa_val(pos)
                if pos:
                    return pos
                if i + 1 < len(linhas):
                    prox = limpa_val(linhas[i + 1])
                    if prox:
                        return prox
    return None


re_nota = r"Nota\s*N[º°o]?\s*[:#-]?\s*(?:\r?\n|\s)*.*?(\d{6,})"
re_desc = r"DESCRIÇÃO DOS SERVIÇOS\s*(.+?)\s*(?:C[ÓO]DIGO\s+DA\s+ATIVIDADE|INFORMA|TRIBUTOS|VALORES\s+DO\s+PRESTADOR)"
re_iss_valor  = r"(?:\(\=\)\s*)?Valor\s+do\s+ISS[^\d]*?(\d{1,3}(?:\.\d{3})*,\d{2})"
re_iss_retido = r"ISS\s+Retido[^\d]*?(\d{1,3}(?:\.\d{3})*,\d{2})"

def extrair_valor_dos_servicos(pdf_path: str) -> str | None:
    pat_num = re.compile(r"\d{1,3}(?:\.\d{3})*,\d{2}")
    alvo = re.compile(r"\bValor\s+dos\s+Servi[cç]os\b", re.I)

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        left_bbox = (0, 0, page.width * 0.55, page.height)
        left = page.crop(left_bbox)

        words = left.extract_words(x_tolerance=3.0, y_tolerance=6)
        if not words:
            return None

        rows = {}
        for w in words:
            ymid = round((w["top"] + w["bottom"]) / 2)
            rows.setdefault(ymid, []).append(w)

        for y in sorted(rows.keys()):
            linha = sorted(rows[y], key=lambda w: w["x0"])
            joined = " ".join(w["text"] for w in linha)

            if re.search(r"Dedução", joined, flags=re.I):
                continue

            if not alvo.search(joined):
                continue

            acc = ""
            x_anchor = None
            for w in linha:
                acc = (acc + " " + w["text"]) if acc else w["text"]
                if alvo.search(acc):
                    x_anchor = w["x1"]
            if x_anchor is None:
                continue

            direita_text = " ".join(w["text"] for w in linha if w["x0"] > x_anchor - 0.5)
            m = pat_num.search(direita_text)
            if m:
                return m.group(0)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            txt = page.extract_text(x_tolerance=2.5, y_tolerance=4) or ""
    except Exception:
        txt = ""

    if txt:
        for raw in txt.splitlines():
            line = norm(raw)
            if re.search(r"Valor\s+dos\s+Servi[cç]os", line, flags=re.I) and not re.search(r"Dedução", line, flags=re.I):
                m = re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}", line)
                if m:
                    return m.group(0)

    return None


files = [f for f in os.listdir(directory) if f.lower().endswith(".pdf")]
if not files:
    raise Exception("Não existem arquivos PDF no diretório")

wb = Workbook()
ws = wb.active
ws.title = "notas importadas"
ws["A1"] = "Nota Fiscal"
ws["B1"] = "Razão Social"
ws["C1"] = "Descrição"
ws["D1"] = "Valor dos Serviços"
ws["E1"] = "ISS (Valor)"
ws["F1"] = "ISS Retido"

row = 2
for file in files:
    fp = os.path.join(directory, file)
    with pdfplumber.open(fp) as pdf:
        page = pdf.pages[0]
        txt = page.extract_text(x_tolerance=2.5, y_tolerance=4) or ""

    m_nota = re.search(re_nota, txt, flags=re.I | re.DOTALL)
    if not m_nota:
        top_lines = [l.strip() for l in txt.splitlines() if l.strip()]
        topo = "\n".join(top_lines[:12])
        m_nota = re.search(r"\b(\d{6,})\b", topo)
    nota = m_nota.group(1) if m_nota else ""

    razao = extrair_razao_posicional(fp) or ""

    m_desc = re.search(re_desc, txt, flags=re.I | re.DOTALL)
    desc = norm(re.sub(r"\s*\n\s*", " ", m_desc.group(1))) if m_desc else ""

    m_iss = re.search(re_iss_valor, txt, flags=re.I)
    iss_valor = m_iss.group(1) if m_iss else ""

    valor_servicos = extrair_valor_dos_servicos(fp) or ""

    m_iss_ret = re.search(re_iss_retido, txt, flags=re.I)
    iss_retido = m_iss_ret.group(1) if m_iss_ret else ""

    ws[f"A{row}"] = nota
    ws[f"B{row}"] = razao
    ws[f"C{row}"] = desc
    ws[f"D{row}"] = valor_servicos
    ws[f"E{row}"] = iss_valor
    ws[f"F{row}"] = iss_retido
    row += 1

ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
out_path = fr"{directory}\Invoices - {ts}.xlsx"
wb.save(out_path)

try:
    subprocess.Popen([soffice_path, out_path], shell=False)
except Exception:
    try:
        os.startfile(out_path)
    except Exception:
        pass

print(f"Planilha salva em: {out_path}")
